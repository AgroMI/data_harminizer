from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

FIXTURES_DIR = Path(__file__).resolve().parent


def _save_workbook(workbook: Workbook, filename: str) -> None:
    target = FIXTURES_DIR / filename
    workbook.save(target)
    workbook.close()


def build_simple_semantic_fixture() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "FieldData"

    rows = [
        ["Date", "Plot ID", "Variety", "Treatment", "Yield t/ha", "Moisture", "Plant Height m", "Notes"],
        ["2026-05-01", "P1", "Apex", "control", 12.5, 17.2, 1.12, "north edge"],
        ["2026-05-02", "P2", "Apex", "treated", 13.1, 16.8, 1.08, "manual check"],
        ["2026-05-03", "P3", "Nova", "control", 12.9, 17.5, 1.15, "ok"],
        ["2026-05-04", "P4", "Nova", "treated", 30.4, 16.4, 1.11, "outlier candidate"],
    ]
    for row in rows:
        sheet.append(row)

    _save_workbook(workbook, "simple_semantic_fixture.xlsx")


def build_multi_sheet_fixture() -> None:
    workbook = Workbook()
    yield_sheet = workbook.active
    yield_sheet.title = "Yield2026"

    for row in [
        ["Date", "Plot ID", "Variety", "Treatment", "Yield", "Notes"],
        ["2026-05-10", "Y1", "Apex", "control", 10.8, "ok"],
        ["2026-05-11", "Y2", "Apex", "treated", 11.4, "ok"],
        ["2026-05-12", "Y3", "Nova", "control", 10.9, "ok"],
        ["2026-05-13", "Y4", "Nova", "treated", 11.7, "ok"],
    ]:
        yield_sheet.append(row)

    moisture_sheet = workbook.create_sheet("Moisture2026")
    for row in [
        ["Date", "Plot ID", "Location", "Moisture"],
        ["2026-05-10", "M1", "north", 18.1],
        ["2026-05-11", "M2", "south", 17.7],
        ["2026-05-12", "M3", "north", 18.3],
        ["2026-05-13", "M4", "south", 17.5],
    ]:
        moisture_sheet.append(row)

    _save_workbook(workbook, "multi_sheet_fixture.xlsx")


def build_noisy_fixture() -> None:
    workbook = Workbook()
    measurements = workbook.active
    measurements.title = "Measurements"

    for row in [
        [None, None, "Harvest metrics", None, None],
        ["Date", "Plot ID", "Yield", "Moisture", "Notes"],
        [None, None, "kg_ha", "pct", None],
        ["2026-06-01", "P10", 14.1, 18.4, "ok"],
        ["2026-06-01", "P10", None, 17.9, "duplicate and missing yield"],
        ["2026-06-03", None, 13.8, None, "missing dimension"],
        ["not-a-date", "P13", None, 18.1, "missing date"],
    ]:
        measurements.append(row)

    notes = workbook.create_sheet("Notes")
    notes["A1"] = "Metadata only"

    _save_workbook(workbook, "noisy_fixture.xlsx")


def main() -> None:
    FIXTURES_DIR.mkdir(parents=True, exist_ok=True)
    build_simple_semantic_fixture()
    build_multi_sheet_fixture()
    build_noisy_fixture()


if __name__ == "__main__":
    main()
