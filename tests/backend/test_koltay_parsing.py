"""Regression tests for agrarian Excel parsing with real-world trial fixtures."""

from __future__ import annotations

from pathlib import Path
import sys
import types

import pytest
from openpyxl import Workbook

from backend.app.services.preview_service import parse_upload_source
from etl.block_detector import detect_blocks_with_positions
from etl.excel_reader import _apply_merged_ranges, read_excel_workbook
from etl.preview_schema import (
    _row_is_header_candidate,
    _row_is_preamble_candidate,
    build_preview,
    classify_block_semantics,
    extract_table_details_from_preview_block,
    extract_table_from_block_cells,
)
from etl.quality_validation import validate_observation_records

FIXTURES = Path(__file__).resolve().parent.parent / "fixtures"


def _build_staging_rows_for_test(preview):
    psycopg_module = sys.modules.get("psycopg")
    psycopg_types_module = sys.modules.get("psycopg.types")
    psycopg_types_json_module = sys.modules.get("psycopg.types.json")
    fastapi_module = sys.modules.get("fastapi")
    backend_db_module = sys.modules.get("backend.app.db")

    try:
        if psycopg_module is None:
            psycopg = types.ModuleType("psycopg")
            psycopg.Connection = object
            psycopg.Cursor = object
            sys.modules["psycopg"] = psycopg
        if psycopg_types_module is None:
            sys.modules["psycopg.types"] = types.ModuleType("psycopg.types")
        if psycopg_types_json_module is None:
            psycopg_types_json = types.ModuleType("psycopg.types.json")

            class Json:  # pragma: no cover - test shim
                def __init__(self, value):
                    self.value = value

            psycopg_types_json.Json = Json
            sys.modules["psycopg.types.json"] = psycopg_types_json
        if fastapi_module is None:
            fastapi = types.ModuleType("fastapi")

            class HTTPException(Exception):  # pragma: no cover - test shim
                def __init__(self, status_code, detail):
                    super().__init__(detail)
                    self.status_code = status_code
                    self.detail = detail

            fastapi.HTTPException = HTTPException
            sys.modules["fastapi"] = fastapi
        if backend_db_module is None:
            backend_db = types.ModuleType("backend.app.db")
            backend_db.get_conn = lambda: None
            sys.modules["backend.app.db"] = backend_db

        from backend.app.services.uploads.commit_service import build_all_staging_rows

        return build_all_staging_rows(upload_id="demo", blocks=preview["blocks"])
    finally:
        if psycopg_module is None:
            sys.modules.pop("psycopg", None)
        if psycopg_types_module is None:
            sys.modules.pop("psycopg.types", None)
        if psycopg_types_json_module is None:
            sys.modules.pop("psycopg.types.json", None)
        if fastapi_module is None:
            sys.modules.pop("fastapi", None)
        if backend_db_module is None:
            sys.modules.pop("backend.app.db", None)


# ---------------------------------------------------------------------------
# 1. _apply_merged_ranges
# ---------------------------------------------------------------------------

class TestApplyMergedRanges:
    def test_horizontal_merge_propagates_value(self):
        rows = [["Title", None, None, None]]
        _apply_merged_ranges(rows, [(0, 1, 0, 4)])
        assert rows[0] == ["Title", "Title", "Title", "Title"]

    def test_vertical_merge_propagates_value(self):
        rows = [["Variety A"], [None], [None], [None]]
        _apply_merged_ranges(rows, [(0, 4, 0, 1)])
        assert all(row[0] == "Variety A" for row in rows)

    def test_rectangular_merge(self):
        rows = [["X", None], [None, None]]
        _apply_merged_ranges(rows, [(0, 2, 0, 2)])
        assert rows[0] == ["X", "X"]
        assert rows[1] == ["X", "X"]

    def test_out_of_bounds_range_does_not_raise(self):
        rows = [["A", "B"]]
        _apply_merged_ranges(rows, [(5, 8, 0, 2)])  # beyond len(rows)
        assert rows == [["A", "B"]]

    def test_multiple_ranges(self):
        rows = [["N0", None, "N40", None]]
        _apply_merged_ranges(rows, [(0, 1, 0, 2), (0, 1, 2, 4)])
        assert rows[0] == ["N0", "N0", "N40", "N40"]


# ---------------------------------------------------------------------------
# 2. Merged cell expansion via the real Excel reader
# ---------------------------------------------------------------------------

def _make_merged_workbook() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="N 0")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws.cell(row=1, column=3, value="N 40")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    ws.cell(row=2, column=1, value="Variety A")
    ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
    for r, v in zip(range(2, 5), [1.0, 2.0, 3.0]):
        ws.cell(row=r, column=2, value=v)
        ws.cell(row=r, column=3, value=v + 0.5)
        ws.cell(row=r, column=4, value=v + 1.0)
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestMergedCellExpansion:
    def test_horizontal_header_merge_expanded(self):
        sheets = read_excel_workbook(_make_merged_workbook(), filename="test.xlsx")
        row0 = sheets[0]["rows"][0]
        assert row0[0] == "N 0"
        assert row0[1] == "N 0", "merged cell should be propagated to adjacent column"
        assert row0[2] == "N 40"
        assert row0[3] == "N 40"

    def test_vertical_label_merge_expanded(self):
        sheets = read_excel_workbook(_make_merged_workbook(), filename="test.xlsx")
        rows = sheets[0]["rows"]
        assert rows[1][0] == "Variety A"
        assert rows[2][0] == "Variety A", "vertical merge should fill rows 2 and 3"
        assert rows[3][0] == "Variety A"


# ---------------------------------------------------------------------------
# 3. Uniform-merged row detection in preview_schema
# ---------------------------------------------------------------------------

class TestUniformMergedRowClassification:
    def test_full_width_title_row_is_preamble(self):
        row = ["Koltay búza 2007."] * 16
        assert _row_is_preamble_candidate(row)
        assert not _row_is_header_candidate(row)

    def test_full_width_metadata_row_is_preamble(self):
        row = ["Nettó parc.ter.: 9,5 m * 1,44 m = 13,68 m²"] * 8
        assert _row_is_preamble_candidate(row)

    def test_block_header_row_same_value_is_preamble(self):
        row = [None, None, "Parcellatermés kg/parc.", "Parcellatermés kg/parc.",
               "Parcellatermés kg/parc.", "Parcellatermés kg/parc."]
        assert _row_is_preamble_candidate(row)
        assert not _row_is_header_candidate(row)

    def test_diverse_header_row_is_header(self):
        row = ["N 0", "N 0", "N 40", "N 40", "N 80", "N 80"]
        assert _row_is_header_candidate(row)
        assert not _row_is_preamble_candidate(row)

    def test_sub_column_letters_are_header(self):
        row = [None, None, "A", "B", "C", "D", "E", "F"]
        assert _row_is_header_candidate(row)


# ---------------------------------------------------------------------------
# 4. Label-bridge split: sparse group-name columns with unique_ratio == 1.0
# ---------------------------------------------------------------------------

def _build_sparse_group_label_rows(n_groups: int = 2, rows_per_group: int = 15) -> list[list[object]]:
    """Side-by-side tables separated by a column that has one label per group."""
    header = [None, "plot", "I.", "II.", "III.", "IV.", None, "plot", "I.", "II.", "III.", "IV."]
    rows: list[list[object]] = [header]
    groups = [f"Group {chr(65 + i)}" for i in range(n_groups)]
    for g_idx, group in enumerate(groups):
        for p in range(1, rows_per_group + 1):
            left_label = group if p == 1 else None
            right_label = group if p == 1 else None
            rows.append([
                left_label, p,
                5.0 + g_idx + p * 0.01, 5.1 + g_idx + p * 0.01,
                5.2 + g_idx + p * 0.01, 5.3 + g_idx + p * 0.01,
                right_label, p,
                12.0 + g_idx + p * 0.01, 12.1 + g_idx + p * 0.01,
                12.2 + g_idx + p * 0.01, 12.3 + g_idx + p * 0.01,
            ])
    return rows


class TestLabelBridgeSparseGroupNames:
    def test_two_tables_split_with_two_group_labels(self):
        rows = _build_sparse_group_label_rows(n_groups=2, rows_per_group=15)
        blocks = detect_blocks_with_positions(rows)
        assert len(blocks) == 2, (
            f"expected 2 blocks from side-by-side tables with group labels, got {len(blocks)}"
        )

    def test_two_tables_split_with_seven_variety_labels(self):
        rows = _build_sparse_group_label_rows(n_groups=7, rows_per_group=5)
        blocks = detect_blocks_with_positions(rows)
        assert len(blocks) == 2, (
            f"expected 2 blocks with 7-variety group labels, got {len(blocks)}"
        )

    def test_left_and_right_blocks_non_overlapping(self):
        rows = _build_sparse_group_label_rows()
        left, right = detect_blocks_with_positions(rows)
        assert left["col_end"] < right["col_start"]


# ---------------------------------------------------------------------------
# 5. Real Koltay file parsing
# ---------------------------------------------------------------------------

class TestRealKoltay2007:
    """Real Koltay búza 2007.xls – variety × N-level field trial."""

    @pytest.fixture(scope="class")
    def blocks(self):
        data = (FIXTURES / "Koltay búza 2007.xls").read_bytes()
        sheets = read_excel_workbook(data, filename="Koltay búza 2007.xls")
        return detect_blocks_with_positions(sheets[0]["rows"], merge_vertical_blocks=True)

    def test_four_blocks_detected(self, blocks):
        assert len(blocks) == 4, f"expected 4 blocks, got {len(blocks)}"

    def test_two_main_blocks_side_by_side(self, blocks):
        left, right = blocks[0], blocks[1]
        assert left["col_end"] < right["col_start"]
        assert left["row_start"] == right["row_start"]
        assert left["row_end"] == right["row_end"]

    def test_main_blocks_have_n_level_headers(self, blocks):
        for block in blocks[:2]:
            table = extract_table_from_block_cells(block["rows"])
            headers = table["headers"]
            assert any("n_0" in h for h in headers), f"N0 missing: {headers}"
            assert any("n_40" in h for h in headers), f"N40 missing: {headers}"
            assert any("n_80" in h for h in headers), f"N80 missing: {headers}"

    def test_main_blocks_have_60_numeric_data_rows(self, blocks):
        for i, block in enumerate(blocks[:2]):
            table = extract_table_from_block_cells(block["rows"])
            numeric_headers = [h for h in table["headers"] if "n_" in h]
            numeric_rows = [
                r for r in table["data_rows"]
                if any(isinstance(r.get(h), (int, float)) for h in numeric_headers)
            ]
            assert len(numeric_rows) == 60, (
                f"block {i+1}: expected 60 numeric rows (12 varieties × 5 reps), "
                f"got {len(numeric_rows)}"
            )

    def test_main_blocks_drop_blank_separator_rows(self, blocks):
        for block in blocks[:2]:
            table = extract_table_from_block_cells(block["rows"])
            assert all(any(value is not None for value in row.values()) for row in table["data_rows"])

    def test_main_blocks_preserve_and_fill_variety_labels(self, blocks):
        table = extract_table_from_block_cells(blocks[0]["rows"])
        first_group = table["data_rows"][:5]
        assert len(first_group) == 5
        assert {row["column_1"] for row in first_group} == {"1. Mv Toborzó"}

    def test_preview_inferrs_variety_dimension_for_group_label_column(self, blocks):
        preview = build_preview(
            file_name="Koltay búza 2007.xls",
            block_records=[
                {
                    "block_id": "S1_B1",
                    "sheet_name": "feldolgozás",
                    **blocks[0],
                }
            ],
        )
        suggestions = {
            item["column"]: item
            for item in preview["blocks"][0]["type_suggestions"]
        }
        assert suggestions["column_1"]["semantic_role"] == "dimension"
        assert suggestions["column_1"]["canonical_dimension"] == "variety"


class TestRealKoltay2007DefaultBlockSeparation:
    @pytest.fixture(scope="class")
    def blocks(self):
        data = (FIXTURES / "Koltay búza 2007.xls").read_bytes()
        sheets = read_excel_workbook(data, filename="Koltay búza 2007.xls")
        return detect_blocks_with_positions(sheets[0]["rows"])

    @pytest.fixture(scope="class")
    def preview(self):
        data = (FIXTURES / "Koltay búza 2007.xls").read_bytes()
        return parse_upload_source(data, filename="Koltay búza 2007.xls")["preview"]

    def test_detect_blocks_keeps_vertical_stack_separate_by_default(self, blocks):
        assert len(blocks) == 26, f"expected 26 vertically separated blocks, got {len(blocks)}"

    def test_first_24_detected_blocks_are_small_tables(self, blocks):
        row_counts = [len(extract_table_from_block_cells(block["rows"])["data_rows"]) for block in blocks]
        assert row_counts[:24] == [5] * 24
        assert row_counts[24:] == [12, 12]

    def test_preview_keeps_detected_blocks_as_distinct_entities(self, preview):
        assert preview["block_count"] == 26
        assert not any(block["row_count"] == 74 for block in preview["blocks"])

    def test_preview_does_not_show_a_single_long_normalized_main_table(self, preview):
        sample_row_counts = [len(block["sample_rows"]) for block in preview["blocks"]]
        assert sample_row_counts[:24] == [5] * 24
        assert max(sample_row_counts) == 12

    def test_preview_inherits_shared_headers_for_later_same_span_blocks(self, preview):
        anchor_headers = preview["blocks"][0]["headers"]
        sibling = preview["blocks"][2]

        assert sibling["headers"] == anchor_headers
        assert "n_0_a" in sibling["headers"]
        assert "column_3" not in sibling["headers"]

        suggestion_columns = [item["column"] for item in sibling["type_suggestions"]]
        assert suggestion_columns == anchor_headers

    def test_downstream_table_extraction_uses_inherited_schema(self, preview):
        sibling = preview["blocks"][2]
        table = extract_table_details_from_preview_block(sibling)

        assert table["headers"] == preview["blocks"][0]["headers"]
        assert table["data_rows"]
        assert set(table["data_rows"][0]) == set(preview["blocks"][0]["headers"])

    def test_preview_keeps_sparse_group_label_as_variety_dimension(self, preview):
        for block in preview["blocks"][:24]:
            suggestion = next(item for item in block["type_suggestions"] if item["column"] == "column_1")
            assert suggestion["semantic_role"] == "dimension"
            assert suggestion["canonical_dimension"] == "variety"
            assert "annotation_like" not in suggestion["warnings"]

    def test_preview_classifies_trial_and_summary_blocks_separately(self, preview):
        classifications = [block["semantic_classification"] for block in preview["blocks"]]
        assert classifications[:24] == ["observation_like"] * 24
        assert classifications[24:] == ["summary_like", "summary_like"]

        decisions = [block["commit_decision"] for block in preview["blocks"]]
        assert decisions[:24] == ["commit_observations"] * 24
        assert decisions[24:] == ["skip_summary_block", "skip_summary_block"]

    def test_harmonized_rows_preserve_variety_measure_and_treatment_metadata(self, preview):
        rows = _build_staging_rows_for_test(preview)
        validate_observation_records(rows)

        assert len(rows) == 960
        assert {row["block_id"] for row in rows} == {block["block_id"] for block in preview["blocks"][:24]}
        assert not any(str(row["source_column"]) in {"a", "b", "c", "d", "e", "f", "g", "h"} for row in rows)

        main_rows = [row for row in rows if str(row["source_column"]).startswith("n_")]
        assert len(main_rows) == 960
        assert {row["validation_status"] for row in main_rows} == {"valid"}

        first = main_rows[0]
        assert first["variety"] == "1. Mv Toborzó"
        assert first["variable"] == "yield"
        assert first["treatment"] == "n_0"
        assert first["dimensions_json"]["raw_measure_name"] == "n_0_a"
        assert first["dimensions_json"]["meaning"] == "yield"
        assert first["dimensions_json"]["replicate"] == "I."
        assert first["dimensions_json"]["replicate_column_code"] == "a"

    def test_summary_block_debug_shape_matches_skip_heuristic(self, preview):
        summary_block = preview["blocks"][-1]
        semantics = classify_block_semantics(summary_block)

        assert summary_block["block_id"] == "S1_B26"
        assert semantics["semantic_classification"] == "summary_like"
        assert semantics["commit_decision"] == "skip_summary_block"
        assert semantics["skip_reason"] == "summary_like_block"


class TestRealKoltay2017:
    """Real Koltay búza 2017.xlsx – treatment group × replicate trial."""

    @pytest.fixture(scope="class")
    def blocks(self):
        data = (FIXTURES / "Koltay búza 2017.xlsx").read_bytes()
        sheets = read_excel_workbook(data, filename="Koltay búza 2017.xlsx")
        return detect_blocks_with_positions(sheets[0]["rows"], merge_vertical_blocks=True)

    def test_four_blocks_detected(self, blocks):
        assert len(blocks) == 4, f"expected 4 blocks (measurement types), got {len(blocks)}"

    def test_blocks_ordered_left_to_right(self, blocks):
        starts = [b["col_start"] for b in blocks]
        assert starts == sorted(starts)

    def test_each_block_has_replicate_headers(self, blocks):
        for i, block in enumerate(blocks):
            table = extract_table_from_block_cells(block["rows"])
            rep_headers = [h for h in table["headers"] if h in {"i.", "ii.", "iii.", "iv."}]
            assert len(rep_headers) == 4, (
                f"block {i+1} should have 4 replicate columns, got {rep_headers}"
            )

    def test_each_block_has_120_numeric_rows(self, blocks):
        for i, block in enumerate(blocks):
            table = extract_table_from_block_cells(block["rows"])
            rep_header = next((h for h in table["headers"] if h == "i."), None)
            assert rep_header is not None
            numeric_rows = [
                r for r in table["data_rows"]
                if isinstance(r.get(rep_header), (int, float))
            ]
            assert len(numeric_rows) == 120, (
                f"block {i+1}: expected 120 numeric rows (8 groups × 15 plots), "
                f"got {len(numeric_rows)}"
            )


class TestRealKoltay2024:
    """Real Koltay búza 2024.xlsx – flat harvester log table."""

    @pytest.fixture(scope="class")
    def blocks(self):
        data = (FIXTURES / "Koltay búza 2024.xlsx").read_bytes()
        sheets = read_excel_workbook(data, filename="Koltay búza 2024.xlsx")
        return detect_blocks_with_positions(sheets[0]["rows"])

    def test_single_block_detected(self, blocks):
        assert len(blocks) == 1, f"expected 1 block (flat table), got {len(blocks)}"

    def test_block_has_480_data_rows(self, blocks):
        table = extract_table_from_block_cells(blocks[0]["rows"])
        assert len(table["data_rows"]) == 480, (
            f"expected 480 data rows, got {len(table['data_rows'])}"
        )

    def test_block_has_weight_and_moisture_headers(self, blocks):
        table = extract_table_from_block_cells(blocks[0]["rows"])
        headers = table["headers"]
        assert any("súly" in h or "weight" in h.lower() for h in headers), (
            f"weight column missing: {headers}"
        )
        assert any("nedvesség" in h or "moisture" in h.lower() for h in headers), (
            f"moisture column missing: {headers}"
        )
