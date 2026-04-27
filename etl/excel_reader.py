import csv
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Callable

import xlrd
from openpyxl import load_workbook

from etl.types import WorkbookSheet

WorkbookReader = Callable[[bytes], list[WorkbookSheet]]
TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "cp1250", "iso-8859-2", "latin-1")
CSV_DELIMITERS = ",;\t|"


def _normalize_row(row: tuple[Any, ...] | list[Any]) -> list[Any]:
    return list(row)


def _apply_merged_ranges(
    rows: list[list[Any]],
    merged_ranges: list[tuple[int, int, int, int]],
) -> None:
    """Propagate top-left value across each merged range (in-place).

    merged_ranges entries: (row_start, row_end_excl, col_start, col_end_excl), 0-indexed.
    """
    for row_start, row_end_excl, col_start, col_end_excl in merged_ranges:
        if row_start >= len(rows):
            continue
        top_row = rows[row_start]
        top_val = top_row[col_start] if col_start < len(top_row) else None
        for ri in range(row_start, min(row_end_excl, len(rows))):
            for ci in range(col_start, col_end_excl):
                if ci < len(rows[ri]):
                    rows[ri][ci] = top_val


def _read_workbook_openpyxl(file_bytes: bytes) -> list[WorkbookSheet]:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheets: list[WorkbookSheet] = []

    for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
        rows = [_normalize_row(row) for row in worksheet.iter_rows(values_only=True)]
        merged_ranges = [
            (r.min_row - 1, r.max_row, r.min_col - 1, r.max_col)
            for r in worksheet.merged_cells.ranges
        ]
        _apply_merged_ranges(rows, merged_ranges)
        sheets.append(
            {
                "sheet_index": sheet_index,
                "sheet_name": worksheet.title,
                "rows": rows,
            }
        )

    workbook.close()
    return sheets


def _convert_xlrd_cell(cell: xlrd.sheet.Cell, datemode: int) -> Any:
    if cell.ctype == xlrd.XL_CELL_DATE:
        # Keep date/datetime values typed for downstream inference.
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype in {xlrd.XL_CELL_ERROR, xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    return cell.value


def _read_workbook_xlrd(file_bytes: bytes) -> list[WorkbookSheet]:
    workbook = xlrd.open_workbook(file_contents=file_bytes)
    sheets: list[WorkbookSheet] = []

    for sheet_index, sheet in enumerate(workbook.sheets(), start=1):
        rows: list[list[Any]] = []

        for row_idx in range(sheet.nrows):
            rows.append(
                [
                    _convert_xlrd_cell(sheet.cell(row_idx, col_idx), workbook.datemode)
                    for col_idx in range(sheet.ncols)
                ]
            )

        # xlrd merged_cells: list of (rlo, rhi, clo, chi) — rhi and chi are exclusive
        merged_ranges = list(sheet.merged_cells)
        _apply_merged_ranges(rows, merged_ranges)

        sheets.append(
            {
                "sheet_index": sheet_index,
                "sheet_name": sheet.name,
                "rows": rows,
            }
        )

    return sheets


def _ordered_readers(filename: str | None) -> tuple[WorkbookReader, WorkbookReader]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".xls":
        return _read_workbook_xlrd, _read_workbook_openpyxl
    return _read_workbook_openpyxl, _read_workbook_xlrd


def _decode_text_table(file_bytes: bytes) -> str:
    errors: list[UnicodeDecodeError] = []
    for encoding in TEXT_ENCODINGS:
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError as exc:
            errors.append(exc)

    raise RuntimeError("Could not decode text table as a supported text encoding") from (
        errors[-1] if errors else None
    )


def _normalize_text_cell(value: str) -> str | None:
    if value == "":
        return None
    return value


def _sniff_csv_dialect(text: str, filename: str | None) -> csv.Dialect:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".tsv":
        return csv.excel_tab

    sample = text[:8192]
    try:
        return csv.Sniffer().sniff(sample, delimiters=CSV_DELIMITERS)
    except csv.Error:
        return csv.excel


def _read_text_table(file_bytes: bytes, filename: str | None = None) -> list[WorkbookSheet]:
    text = _decode_text_table(file_bytes)
    dialect = _sniff_csv_dialect(text, filename)
    reader = csv.reader(StringIO(text, newline=""), dialect=dialect)
    rows = [[_normalize_text_cell(cell) for cell in row] for row in reader]
    sheet_name = Path(filename or "").stem or "Table"

    return [
        {
            "sheet_index": 1,
            "sheet_name": sheet_name,
            "rows": rows,
        }
    ]


def _is_text_table(filename: str | None) -> bool:
    return Path(filename or "").suffix.lower() in {".csv", ".tsv"}


def read_excel_workbook(file_bytes: bytes, filename: str | None = None) -> list[WorkbookSheet]:
    errors: list[Exception] = []

    for reader in _ordered_readers(filename):
        try:
            return reader(file_bytes)
        except Exception as exc:  # pragma: no cover - fallback chain
            errors.append(exc)

    raise RuntimeError("Could not parse Excel workbook as .xlsx or .xls") from (
        errors[-1] if errors else None
    )


def read_tabular_source(file_bytes: bytes, filename: str | None = None) -> list[WorkbookSheet]:
    if _is_text_table(filename):
        return _read_text_table(file_bytes, filename=filename)

    return read_excel_workbook(file_bytes=file_bytes, filename=filename)
